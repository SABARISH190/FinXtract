from unittest import result
import sys
import asyncio

if sys.platform.startswith("win"):
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
from playwright.sync_api import sync_playwright
import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
import base64
from io import BytesIO, StringIO
import os
# ------------------------------
# Helper: background image
# ------------------------------
def load_bg_base64(image_path):
    with open(image_path, "rb") as f:
        data = f.read()
    return base64.b64encode(data).decode()


# ------------------------------
# Screener search by company name
# ------------------------------
def find_screener_company_by_name(company_name):

    url = "https://www.screener.in/api/company/search/"

    r = requests.get(
        url,
        params={"q": company_name},
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=15
    )
    r.raise_for_status()

    data = r.json()

    if not data:
        return None

    return "https://www.screener.in" + data[0]["url"]

def get_screener_html_with_expanded_rows(url):

    print("\n--- Opening page:", url)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        page.goto(url, timeout=60000)
        page.wait_for_selector(
            "//h2[normalize-space()='Peer comparison']/following::table[1]",
            timeout=60000
        )

        # give JS time to update CMP / P-E cells
        page.wait_for_timeout(2500)

        def expand_all_tables():

            tables = page.locator(
                "//h2/following::table[1] | //h3/following::table[1]"
            )

            for t in range(tables.count()):
                table = tables.nth(t)

                while True:

                    # 👉 IMPORTANT: handle BOTH + and >
                    buttons = table.locator("""
                        xpath=.//button[
                            .//span[contains(@class,'blue-icon')
                            and (normalize-space(text())='+' or normalize-space(text())='>')]
                        ]
                    """)

                    if buttons.count() == 0:
                        break

                    btn = buttons.first
                    before = table.locator("tr").count()

                    try:
                        btn.scroll_into_view_if_needed()
                        btn.click(force=True, timeout=3000)
                        page.wait_for_timeout(250)
                    except:
                        break

                    after = table.locator("tr").count()

                    # nothing expanded -> remove this expander and continue
                    if after <= before:
                        try:
                            btn.evaluate("b => b.remove()")
                        except:
                            break

        # -------------------- Periodic / Quarterly --------------------
        q_btn = page.locator("//button[normalize-space()='Quarterly']")
        if q_btn.count():
            q_btn.first.click(force=True)
            page.wait_for_timeout(600)

        expand_all_tables()
        html_periodic = page.content()

        # -------------------- Yearly --------------------
        y_btn = page.locator("//button[normalize-space()='Yearly']")
        if y_btn.count():
            y_btn.first.click(force=True)
            page.wait_for_timeout(600)

        expand_all_tables()
        html_yearly = page.content()

        browser.close()

        return html_periodic, html_yearly
    
#------------------------------
# Fetch live CMP from NSE (for validation)
#------------------------------
    
def fetch_live_cmp_nse(symbol):
    url = f"https://www.nseindia.com/api/quote-equity?symbol={symbol}"

    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json",
        "Referer": "https://www.nseindia.com/"
    }

    s = requests.Session()
    s.get("https://www.nseindia.com", headers=headers, timeout=10)

    r = s.get(url, headers=headers, timeout=10)
    r.raise_for_status()

    data = r.json()
    return float(data["priceInfo"]["lastPrice"])

#------------------------------
#patch peer comparison table with live CMPs from NSE (best effort, for validation only)
#------------------------------
def patch_peer_comparison_with_live_prices(df):

    # ⚠ You MUST maintain this mapping
    # Screener name  ->  NSE symbol
    symbol_map = {
        "Billionbrains": "BILLIONBR",
        "Motil.Oswal.Fin.": "MOTILALOFS",
        "360 ONE": "360ONE",
        "Angel One": "ANGELONE",
        "Nuvama Wealth": "NUVAMA",
        "IIFL Capital": "IIFLCAPS",
        "Anand Rathi Shar": "ANANDRATHI",
        "Emk.Global Fin.": "EMKAY"
    }

    if "Name" not in df.columns or "CMP Rs." not in df.columns:
        return df

    for i in df.index:

        name = str(df.at[i, "Name"]).strip()

        symbol = symbol_map.get(name)
        if not symbol:
            continue

        try:
            live_cmp = fetch_live_cmp_nse(symbol)
            df.at[i, "CMP Rs."] = live_cmp

            # ---- recompute P/E if EPS present
            if "P/E" in df.columns:
                # we cannot recompute without EPS
                # so only overwrite CMP here
                pass

        except Exception:
            pass

    return df

#------------------------------
# Extract bold rows (used for better section naming in some cases)
#------------------------------
def extract_bold_rows(table_soup):
    bold_rows = set()

    rows = table_soup.find_all("tr")
    for i, tr in enumerate(rows):
        first_cell = tr.find(["th", "td"])
        if not first_cell:
            continue

        if first_cell.find(["b", "strong"]):
            bold_rows.add(i)

    return bold_rows

# ------------------------------
# Scrape Screener tables by company name
# ------------------------------
def scrape_screener_financials_by_name(company_name, mode):

    company_url = find_screener_company_by_name(company_name)

    if not company_url:
        return None, {}

    # ---------------------------------
    # ALWAYS normalize base company URL
    # ---------------------------------
    company_url = company_url.replace("/consolidated/", "").rstrip("/") + "/"

    # ---------------------------------
    # Consolidated / Standalone switch
    # ---------------------------------
    if mode == "Consolidated":
        company_url = company_url + "consolidated/"

    # ✅ only ONE html now
    html_periodic, html_yearly = get_screener_html_with_expanded_rows(company_url)

    soup_main = BeautifulSoup(html_periodic, "lxml")
    soup_yearly = BeautifulSoup(html_yearly, "lxml")


    tables = []

    # ---------- take everything except shareholding from periodic snapshot
    for table in soup_main.find_all("table"):

        name = table.find_previous(["h2", "h3"])
        if name and "shareholding" in name.get_text(strip=True).lower():
            continue

        tables.append(table)

    # ---------- take shareholding only from periodic snapshot
    for table in soup_main.find_all("table"):

        name = table.find_previous(["h2", "h3"])
        if not name:
            continue

        if "shareholding" in name.get_text(strip=True).lower():
            tables.append(table)

    # ---------- take shareholding only from yearly snapshot
    for table in soup_yearly.find_all("table"):

        name = table.find_previous(["h2", "h3"])
        if not name:
            continue

        if "shareholding" in name.get_text(strip=True).lower():
            tables.append(table)


    result = {}

    for table in tables:

        try:
            bold_rows = extract_bold_rows(table)
            df = pd.read_html(StringIO(str(table)))[0]
            # attach bold info
            df.attrs["bold_rows"] = bold_rows
        except Exception:
            continue

        # -----------------------------
        # Section name
        # -----------------------------
        name = table.find_previous(["h2", "h3"])

        if name:
            key = name.get_text(strip=True)
        else:
            key = "Table"

        # -----------------------------
        # Fix Raw PDF links
        # -----------------------------
        try:
            rows = table.find_all("tr")

            for tr in rows:

                cells = tr.find_all(["th", "td"])
                if not cells:
                    continue

                first_cell_text = cells[0].get_text(strip=True)

                if first_cell_text.lower() == "raw pdf":

                    links = []
                    for td in cells[1:]:
                        a = td.find("a")
                        if a and a.get("href"):
                            href = a["href"]
                            if href.startswith("/"):
                                href = "https://www.screener.in" + href
                            links.append(href)
                        else:
                            links.append(None)

                    mask = (
                        df.iloc[:, 0]
                        .astype(str)
                        .str.strip()
                        .str.lower()
                        == "raw pdf"
                    )

                    if mask.any():
                        row_index = df[mask].index[0]

                        for i, link in enumerate(links):
                            if i + 1 < len(df.columns):
                                df.iat[row_index, i + 1] = link
        except:
            pass

        # ------------------------------------------------
        # SMART rename for mini KPI tables
        # ------------------------------------------------
        try:
            if df.shape[1] == 2 and df.shape[0] <= 6:

                first_col = str(df.columns[0]).strip().lower()

                if "unnamed" in first_col:
                    label = str(df.iloc[0, 0]).strip()
                else:
                    label = str(df.columns[0]).strip()

                kpi_labels = {
                    "compounded sales growth",
                    "compounded profit growth",
                    "stock price cagr",
                    "return on equity"
                }

                if label.lower() in kpi_labels:
                    key = label
        except:
            pass

        # ---------------------------------------
        # Shareholding pattern naming
        # ---------------------------------------
        if key.lower().startswith("shareholding"):

            try:
                cols = list(df.columns)[1:]

                months = []
                for c in cols:
                    c = str(c).strip()
                    parts = c.split()
                    if parts:
                        months.append(parts[0].lower())

                non_mar = [m for m in months if m != "mar"]

                if len(non_mar) >= 3:
                    key = "Shareholding Pattern (Periodic)"
                else:
                    key = "Shareholding Pattern (Yearly)"

            except:
                pass

        # ------------------------------------------------
        # 🚨 very important: do NOT allow duplicate
        # Shareholding tables
        # ------------------------------------------------
        if key.startswith("Shareholding"):
            if key in result:
                continue

        # -----------------------------
        # Make key unique (others only)
        # -----------------------------
        base_key = key
        cnt = 1
        while key in result:
            cnt += 1
            key = f"{base_key} ({cnt})"

        if key.strip().lower() == "peer comparison":
            df = patch_peer_comparison_with_live_prices(df)

        result[key] = df


    return company_url, result

# ------------------------------
# Layout / health validation
# ------------------------------
def validate_core_sections(tables: dict):

    keys = [k.lower() for k in tables.keys()]

    checks = {
        "Profit & Loss": any("profit" in k and "loss" in k for k in keys),
        "Balance Sheet": any("balance" in k for k in keys),
        "Quarterly Results": any("quarter" in k for k in keys),
    }

    missing = [name for name, ok in checks.items() if not ok]
    return missing


# ------------------------------
# Excel download helper
# ------------------------------
def to_excel_bytes(dfs: dict):

    from openpyxl.styles import Font

    buf = BytesIO()
    writer = pd.ExcelWriter(buf, engine="openpyxl")

    for sheet, df in dfs.items():

        safe_sheet = sheet[:31]

        df.to_excel(writer, sheet_name=safe_sheet, index=False)

        ws = writer.book[safe_sheet]

        # -----------------------------
        # Find "Raw PDF" row
        # -----------------------------
        raw_pdf_row = None

        for r in range(2, ws.max_row + 1):   # skip header
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip().lower() == "raw pdf":
                raw_pdf_row = r
                break

        # -----------------------------
        # Convert URLs into hyperlinks
        # -----------------------------
        if raw_pdf_row is not None:

            for c in range(2, ws.max_column + 1):

                cell = ws.cell(row=raw_pdf_row, column=c)

                if isinstance(cell.value, str) and cell.value.startswith("http"):

                    link = cell.value

                    cell.value = "Raw PDF"
                    cell.hyperlink = link
                    cell.font = Font(color="0563C1", underline="single")

    writer.close()
    buf.seek(0)
    return buf





# ------------------------------
# Session state
# ------------------------------
if "screener_tables" not in st.session_state:
    st.session_state.screener_tables = None

if "screener_company_url" not in st.session_state:
    st.session_state.screener_company_url = None

if "screener_company_name" not in st.session_state:
    st.session_state.screener_company_name = None

if "missing_sections" not in st.session_state:
    st.session_state.missing_sections = []
if "statement_mode" not in st.session_state:
    st.session_state.statement_mode = None
if "fetched" not in st.session_state:
    st.session_state.fetched = False


st.title("FinXtract • Screener Data")

# ------------------------------
# FORM
# ------------------------------
with st.form("fetch_form"):
    company_input = st.text_input("Enter company name (as in Screener)")

    mode = st.radio(
        "Statement type",
        ["Consolidated", "Standalone"],
        horizontal=True
    )

    submit = st.form_submit_button("🚀 Fetch Financials")

# ------------------------------
# Fetch logic
# ------------------------------
if submit:

    if not company_input.strip():
        st.warning("Please enter a company name.")
    else:

        with st.spinner("Searching Screener and fetching financial tables..."):

            try:
                company_url, all_tables = scrape_screener_financials_by_name(
                    company_input.strip(), 
                    mode)


                st.session_state.screener_tables = all_tables
                st.session_state.screener_company_url = company_url
                st.session_state.screener_company_name = company_input.strip()
                st.session_state.statement_mode = mode
                st.session_state.fetched = True



                if all_tables:
                    st.session_state.missing_sections = validate_core_sections(all_tables)
                else:
                    st.session_state.missing_sections = []

            except Exception as e:
                st.error(str(e))
                st.session_state.screener_tables = None
                st.session_state.screener_company_url = None
                st.session_state.screener_company_name = None
                st.session_state.missing_sections = []

# ------------------------------
# Streamlit UI
# ------------------------------
st.set_page_config(page_title="FinXtract (Screener)", layout="wide")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
bg_base64 = load_bg_base64(
    os.path.join(BASE_DIR, "bg", "bg-image.png")
)

if not st.session_state.get("screener_tables", False):

    # ---- First screen (with image)
    st.markdown(f"""
    <style>
    .stApp {{
        background:
            linear-gradient(
                rgba(0, 0, 0, 0.45),
                rgba(0, 0, 0, 0.45)
            ),
            url("data:image/png;base64,{bg_base64}");
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
        background-attachment: fixed;
    }}

    h1,h2,h3,h4,p,label {{
        color: white !important;
    }}

    .stButton > button {{
        background: rgba(0,0,0,0.55);
        color: white;
        border: 1px solid rgba(255,255,255,0.25);
        border-radius: 10px;
    }}

    .screener-link {{
        font-size: 15px;
        font-weight: 600;
    }}

    .screener-link a {{
        color: #ffffff !important;
        text-decoration: underline;
    }}

    /* -----------------------------
    Professional clean table (no background)
    ----------------------------- */

    .fin-table table{{
        width:100%;
        border-collapse:collapse;
        background: transparent !important;
    }}

    .fin-table th{{
        background: transparent !important;
        color:#ffffff !important;
        font-weight:600;
        border-bottom:1px solid rgba(255,255,255,0.25) !important;
    }}

    .fin-table td{{
        background: transparent !important;
        color:#e5e7eb !important;
    }}

    .fin-table th,
    .fin-table td{{
        padding:8px 10px;
        border-right:1px solid rgba(255,255,255,0.08);
        border-bottom:1px solid rgba(255,255,255,0.08);
        font-size:13px;
        white-space:nowrap;
    }}

    /* left border */
    .fin-table th:first-child,
    .fin-table td:first-child{{
        border-left:1px solid rgba(255,255,255,0.08);
    }}

    /* subtle row hover (very light, pro look) */
    .fin-table tr:hover td{{
        background: rgba(255,255,255,0.03) !important;
    }}

    /* links */
    .fin-table a{{
        color:#60a5fa;
        text-decoration:underline;
    }}
    /* inputs */
    div[data-baseweb="input"] > div {{
        background: rgba(15, 18, 25, 0.92) !important;
        backdrop-filter: blur(6px);
        -webkit-backdrop-filter: blur(6px);
        border-radius: 8px !important;
        border: 1px solid rgba(255,255,255,0.08) !important;
        min-height: 44px;
    }}

    div[data-baseweb="input"] > div > div {{
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
    }}

    div[data-baseweb="input"] input {{
        background: transparent !important;
        color: #e6e6e6 !important;
        font-size: 14px;
        padding: 10px 12px !important;
    }}
    /* -----------------------------
    Professional clean table (no background)
    ----------------------------- */

    .fin-table table{{
        width:100%;
        border-collapse:collapse;
        background: transparent !important;
    }}

    .fin-table th{{
        background: transparent !important;
        color:#ffffff !important;
        font-weight:600;
        border-bottom:1px solid rgba(255,255,255,0.25) !important;
    }}

    .fin-table td{{
        background: transparent !important;
        color:#e5e7eb !important;
    }}

    .fin-table th,
    .fin-table td{{
        padding:8px 10px;
        border-right:1px solid rgba(255,255,255,0.08);
        border-bottom:1px solid rgba(255,255,255,0.08);
        font-size:13px;
        white-space:nowrap;
    }}

    /* left border */
    .fin-table th:first-child,
    .fin-table td:first-child{{
        border-left:1px solid rgba(255,255,255,0.08);
    }}

    /* subtle row hover (very light, pro look) */
    .fin-table tr:hover td{{
        background: rgba(255,255,255,0.03) !important;
    }}
    </style>
    """, unsafe_allow_html=True)

if not st.session_state.get("screener_tables", False):

    # ---- First screen (with image)
    st.markdown(f"""
    <style>
    .stApp {{
        background:
            linear-gradient(
                rgba(0, 0, 0, 0.45),
                rgba(0, 0, 0, 0.45)
            ),
            url("data:image/png;base64,{bg_base64}");
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
        background-attachment: fixed;
    }}

    h1,h2,h3,h4,p,label {{
        color: white !important;
    }}

    .stButton > button {{
        background: rgba(0,0,0,0.55);
        color: white;
        border: 1px solid rgba(255,255,255,0.25);
        border-radius: 10px;
    }}

    .screener-link {{
        font-size: 15px;
        font-weight: 600;
    }}

    .screener-link a {{
        color: #ffffff !important;
        text-decoration: underline;
    }}

    /* -----------------------------
    Professional clean table (no background)
    ----------------------------- */

    .fin-table table{{
        width:100%;
        border-collapse:collapse;
        background: transparent !important;
    }}

    .fin-table th{{
        background: transparent !important;
        color:#ffffff !important;
        font-weight:600;
        border-bottom:1px solid rgba(255,255,255,0.25) !important;
    }}

    .fin-table td{{
        background: transparent !important;
        color:#e5e7eb !important;
    }}

    .fin-table th,
    .fin-table td{{
        padding:8px 10px;
        border-right:1px solid rgba(255,255,255,0.08);
        border-bottom:1px solid rgba(255,255,255,0.08);
        font-size:13px;
        white-space:nowrap;
    }}

    /* left border */
    .fin-table th:first-child,
    .fin-table td:first-child{{
        border-left:1px solid rgba(255,255,255,0.08);
    }}

    /* subtle row hover (very light, pro look) */
    .fin-table tr:hover td{{
        background: rgba(255,255,255,0.03) !important;
    }}

    /* links */
    .fin-table a{{
        color:#60a5fa;
        text-decoration:underline;
    }}
    /* inputs */
    div[data-baseweb="input"] > div {{
        background: rgba(15, 18, 25, 0.92) !important;
        backdrop-filter: blur(6px);
        -webkit-backdrop-filter: blur(6px);
        border-radius: 8px !important;
        border: 1px solid rgba(255,255,255,0.08) !important;
        min-height: 44px;
    }}

    div[data-baseweb="input"] > div > div {{
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
    }}

    div[data-baseweb="input"] input {{
        background: transparent !important;
        color: #e6e6e6 !important;
        font-size: 14px;
        padding: 10px 12px !important;
    }}
    /* -----------------------------
    Professional clean table (no background)
    ----------------------------- */

    .fin-table table{{
        width:100%;
        border-collapse:collapse;
        background: transparent !important;
    }}

    .fin-table th{{
        background: transparent !important;
        color:#ffffff !important;
        font-weight:600;
        border-bottom:1px solid rgba(255,255,255,0.25) !important;
    }}

    .fin-table td{{
        background: transparent !important;
        color:#e5e7eb !important;
    }}

    .fin-table th,
    .fin-table td{{
        padding:8px 10px;
        border-right:1px solid rgba(255,255,255,0.08);
        border-bottom:1px solid rgba(255,255,255,0.08);
        font-size:13px;
        white-space:nowrap;
    }}

    /* left border */
    .fin-table th:first-child,
    .fin-table td:first-child{{
        border-left:1px solid rgba(255,255,255,0.08);
    }}

    /* subtle row hover (very light, pro look) */
    .fin-table tr:hover td{{
        background: rgba(255,255,255,0.03) !important;
    }}
    </style>
    """, unsafe_allow_html=True)

else:

    st.markdown("""
    <style>

    .stApp {
        background: #0e1117 !important;
        background-image: none !important;
        background-attachment: scroll !important;
    }

    .stAppViewContainer {
        background: #0e1117 !important;
        background-image: none !important;
    }

    section.main > div {
        background: #0e1117 !important;
        background-image: none !important;
    }

    header, footer {
        background: #0e1117 !important;
    }

    h1,h2,h3,h4,p,label {
        color: #ffffff !important;
    }

    </style>
    """, unsafe_allow_html=True)



# ------------------------------
# Display section
# ------------------------------
tables = st.session_state.screener_tables
company_url = st.session_state.screener_company_url
company_name = st.session_state.screener_company_name
missing = st.session_state.missing_sections
statement_mode = st.session_state.statement_mode


if tables:

    if missing:
        st.warning(
            "⚠ Possible Screener layout change detected. "
            f"Missing core sections: {', '.join(missing)}"
        )

    if company_url:
        st.markdown(
            f"""
            <div class="screener-link">
                🔗 Screener page :
                <a href="{company_url}" target="_blank">{company_url}</a>
            </div>
            """,
            unsafe_allow_html=True
        )

    for name, df in tables.items():
        st.subheader(name)

        df_show = df.copy()

        def make_link(v):
            if isinstance(v, str) and v.startswith("http"):
                return f'<a href="{v}" target="_blank">PDF</a>'
            return v
        df_show=df_show.map(make_link)
        bold_rows = df.attrs.get("bold_rows", set())
        html = df_show.to_html(index=False, escape=False)
        # highlight bold rows
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, "lxml")
        trs = soup.find_all("tr")[1:]  # skip header
        for i, tr in enumerate(trs):
            if i in bold_rows:
                tr["style"] = "font-weight:700; background: rgba(255,255,255,0.03);"
        st.markdown(
            f"<div class='fin-table'>{str(soup)}</div>",
            unsafe_allow_html=True
        )


    excel_buf = to_excel_bytes(tables)

    st.download_button(
        "⬇ Download All Financials as Excel",
        data=excel_buf,
        file_name=f"{company_name.replace(' ', '_')}-{statement_mode.lower()}_screener.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



